#!/usr/bin/env python3
"""
gerar_planilha.py — Skill nps-acao-contato
Gera planilha de detratores NPS MetLife × iFood para ação de contato.

Uso:
    python3 skills/nps-acao-contato/scripts/gerar_planilha.py --data-corte DD/MM/YYYY

Saída:
    data/nps-sinistros/reclamacoes_<MMDD>_<YYYY>.xlsx
    data/nps-sinistros/reclamacoes_<MMDD>_<YYYY>_meta.json

Dependências: openpyxl, requests (pré-instaladas no sandbox ToqanClaw)
Autenticação: gerenciada pelo proxy ToqanClaw — sem credenciais hardcoded.
"""

import argparse
import json
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import requests

# ─── CONFIG ───────────────────────────────────────────────────────────────────


SPREADSHEET_ID = "1vy_33ywbjRY9dYxo_e5jOeNSBSlJofrFkGXAr6xKL-Q"
SHEET_NAME = "NPS base full Metlife"
CACHE_PATH = "/workspace/data/nps-painel/nps_data_cache.json"
OUTPUT_DIR = "/workspace/data/nps-sinistros"

KEYWORDS = [
    "recusado", "negado", "indeferido", "não recebi", "nao recebi",
    "demora", "sem resposta", "desistir", "péssimo", "pessimo",
    "horrível", "horrivel", "decepcionante", "abandona",
    "não funciona", "nao funciona", "problema", "injusto", "mentira",
    "enganado", "cancelado", "não pago", "nao pago", "difícil",
    "dificil", "ruim", "prejudic",
]

# Colunas de saída — ordem exata conforme especificação da skill
COLUMNS_OUT = [
    "nota_nps", "sinistro", "distribuido", "cobertura", "status",
    "mot_recusa", "q2_comentario", "q2_motivo", "driver_uuid",
    "nome", "cpf", "telefone",
]

# Índices de colunas na planilha NPS (0-based, após skip do header)
# 0=sinistro, 1=cobertura, 2=status, 3=valor_pago, 4=mot_recusa,
# 5=cidade, 6=uf, 7=saude, 8=pesquisa, 9=distribuido,
# 10=canal_distribuicao, 11=canal_resposta, 12=respondido,
# 13=q1_nps (nota), 14=q1_comentario, 15=q2_motivo, 16=q2_comentario
# 17=driver_uuid (se presente)
IDX = {
    "sinistro": 0,
    "cobertura": 1,
    "status": 2,
    "mot_recusa": 4,
    "distribuido": 9,
    "nota_nps": 13,
    "q1_comentario": 14,
    "q2_motivo": 15,
    "q2_comentario": 16,
    "driver_uuid": 17,
}

COL_WIDTHS = {
    "nota_nps": 10,
    "sinistro": 12,
    "distribuido": 14,
    "cobertura": 14,
    "status": 14,
    "mot_recusa": 55,
    "q2_comentario": 60,
    "q2_motivo": 30,
    "driver_uuid": 38,
    "nome": 35,
    "cpf": 14,
    "telefone": 18,
}



# ─── HELPERS ──────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Gera planilha de detratores NPS para ação de contato"
    )
    parser.add_argument(
        "--data-corte",
        required=True,
        metavar="DD/MM/YYYY",
        help="Data de corte (inclusive). Formato DD/MM/YYYY.",
    )
    return parser.parse_args()



def parse_date(s):
    """Parse DD/MM/YYYY, YYYY-MM-DD ou M/D/YYYY."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def safe(row, idx, default=""):
    """Retorna elemento de lista por índice, com fallback."""
    if idx < len(row):
        v = row[idx]
        if v is None:
            return default
        return str(v).strip()
    return default


def has_keyword(text):
    """Verifica se o texto contém ao menos uma keyword negativa."""
    if not text:
        return False
    t = text.lower()
    return any(kw in t for kw in KEYWORDS)


# ─── STEP 1: CARREGAR DADOS NPS ────────────────────────────────────────────────


def load_nps_from_sheets():
    """
    Lê a planilha NPS via Google Sheets API.
    O proxy ToqanClaw injeta a autenticação automaticamente.
    Retorna lista de linhas (sem header) ou None se falhar.
    """
    url = (
        f"https://sheets.googleapis.com/v4/spreadsheets/{SPREADSHEET_ID}"
        f"/values/{requests.utils.quote(SHEET_NAME)}?majorDimension=ROWS"
    )
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            rows = data.get("values", [])
            if rows:
                return rows[1:]  # pula header
        print(f"  [Sheets] status {resp.status_code}: {resp.text[:200]}")
    except Exception as exc:
        print(f"  [Sheets] erro de conexão: {exc}")
    return None


def load_nps_from_cache():
    """Fallback: lê dados do cache local JSON."""
    if not os.path.exists(CACHE_PATH):
        return None
    try:
        with open(CACHE_PATH, encoding="utf-8") as f:
            cache = json.load(f)
        rows = cache.get("values", [])
        return rows[1:] if rows else None  # pula header
    except Exception as exc:
        print(f"  [Cache] erro ao ler: {exc}")
    return None


# ─── STEP 2: FILTRAR DETRATORES ────────────────────────────────────────────────


def filter_detractors(all_rows, cutoff_dt):
    """
    Aplica critérios de filtro:
    1. distribuido >= cutoff_dt
    2. nota_nps entre 0 e 6
    3. tem comentário (q1_comentario ou q2_comentario)
    4. comentário tem keyword negativa (com fallback se nenhum match)
    Retorna (lista_final, keyword_filter_applied).
    """
    with_comment = []
    for row in all_rows:
        # Garante comprimento mínimo
        while len(row) < 17:
            row.append("")

        dt = parse_date(safe(row, IDX["distribuido"]))
        if not dt or dt < cutoff_dt:
            continue

        try:
            nota = int(safe(row, IDX["nota_nps"]))
        except (ValueError, TypeError):
            continue

        if nota > 6:
            continue

        q1 = safe(row, IDX["q1_comentario"])
        q2 = safe(row, IDX["q2_comentario"])

        if not q1 and not q2:
            continue

        with_comment.append(row)


    print(f"  Detratores com comentário >= data de corte: {len(with_comment)}")

    with_keyword = [
        r for r in with_comment
        if has_keyword(safe(r, IDX["q1_comentario"]))
        or has_keyword(safe(r, IDX["q2_comentario"]))
    ]
    print(f"  Com keyword de reclamação: {len(with_keyword)}")

    if with_keyword:
        print("  Filtro por keyword aplicado.")
        return with_keyword, True
    else:
        print("  [!] Sem match por keyword — fallback: todos detratores com comentário.")
        return with_comment, False


# ─── STEP 3: CONSULTAR DATABRICKS PII ────────────────────────────────────────


def query_pii(driver_uuids):
    """
    Consulta CPF, nome e telefone no Databricks PII via SQL API.
    O proxy ToqanClaw injeta credenciais automaticamente.
    Retorna dict: driver_uuid -> {cpf, nome, telefone}
    """
    if not driver_uuids:
        return {}

    uuids_csv = "', '".join(driver_uuids)
    sql = (
        "SELECT driver_uuid, cpf, full_name AS nome, phone AS telefone "
        "FROM pii.pii_data_raw.driver_data "
        f"WHERE driver_uuid IN ('{uuids_csv}')"
    )

    pii_map = {}
    try:
        resp = requests.post(
            "https://ifood-prod-main.cloud.databricks.com/api/2.0/sql/statements",
            json={"statement": sql, "warehouse_id": "auto", "wait_timeout": "30s"},
            timeout=60,
        )
        print(f"  [Databricks PII] status: {resp.status_code}")
        if resp.status_code == 200:
            data = resp.json()
            schema = data.get("manifest", {}).get("schema", {}).get("columns", [])
            col_names = [c["name"] for c in schema]
            rows = data.get("result", {}).get("data_array", [])
            for r in rows:
                d = dict(zip(col_names, r))
                uuid_key = d.get("driver_uuid", "")
                if uuid_key:
                    pii_map[uuid_key] = {
                        "cpf": str(d.get("cpf") or "").strip(),
                        "nome": str(d.get("nome") or "").strip(),
                        "telefone": str(d.get("telefone") or "").strip(),
                    }
            print(f"  [Databricks PII] registros encontrados: {len(pii_map)}")
        else:
            print(f"  [Databricks PII] falha: {resp.text[:300]}")
    except Exception as exc:
        print(f"  [Databricks PII] erro: {exc}")

    return pii_map


# ─── STEP 4: MONTAR LINHAS DE SAÍDA ──────────────────────────────────────────

def build_output_rows(final_rows, pii_map):
    """
    Monta lista de dicts com as 12 colunas de saída em ordem exata.
    driver_uuid é lido da coluna 17 da planilha (se existir).
    PII é cruzado pelo driver_uuid.
    """
    output = []
    for row in final_rows:
        while len(row) < 18:
            row.append("")


        nota_raw = safe(row, IDX["nota_nps"])
        try:
            nota_nps = int(nota_raw)
        except (ValueError, TypeError):
            nota_nps = nota_raw

        driver_uuid = safe(row, IDX["driver_uuid"])
        pii = pii_map.get(driver_uuid, {}) if driver_uuid else {}
        cpf_raw = pii.get("cpf", "")
        cpf = "".join(c for c in cpf_raw if c.isdigit()) if cpf_raw else ""


        output.append({
            "nota_nps": nota_nps,
            "sinistro": safe(row, IDX["sinistro"]),
            "distribuido": safe(row, IDX["distribuido"]),
            "cobertura": safe(row, IDX["cobertura"]),
            "status": safe(row, IDX["status"]),
            "mot_recusa": safe(row, IDX["mot_recusa"]),
            "q2_comentario": safe(row, IDX["q2_comentario"]),
            "q2_motivo": safe(row, IDX["q2_motivo"]),
            "driver_uuid": driver_uuid,
            "nome": pii.get("nome", ""),
            "cpf": cpf,
            "telefone": pii.get("telefone", ""),
        })
    return output


# ─── STEP 5: GERAR EXCEL ──────────────────────────────────────────────────────

def write_excel(output_rows, out_path):
    """Gera arquivo Excel com header formatado e linhas de dados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detratores"


    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")


    for col_idx, col_name in enumerate(COLUMNS_OUT, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(col_name, 20)

    for row_idx, row_data in enumerate(output_rows, 2):
        for col_idx, col_name in enumerate(COLUMNS_OUT, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))


    ws.freeze_panes = "A2"


    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


# ─── STEP 6: SALVAR METADADOS ─────────────────────────────────────────────────


def write_meta(output_rows, keyword_filter_applied, meta_path):
    """Salva JSON com metadados da geração."""
    distribuido_dates = [
        parse_date(r.get("distribuido", "")) for r in output_rows
    ]
    distribuido_dates = [d for d in distribuido_dates if d]


    meta = {
        "total_detratores_filtrados": len(output_rows),
        "com_telefone": sum(1 for r in output_rows if r.get("telefone")),
        "data_minima": min(distribuido_dates).strftime("%d/%m/%Y") if distribuido_dates else None,
        "data_maxima": max(distribuido_dates).strftime("%d/%m/%Y") if distribuido_dates else None,
        "gerado_em": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S") + "+00:00",
        "keyword_filter_applied": keyword_filter_applied,
    }

    os.makedirs(os.path.dirname(meta_path), exist_ok=True)
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    return meta


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()

    cutoff_dt = parse_date(args.data_corte)
    if not cutoff_dt:
        print(f"[ERRO] Data de corte inválida: '{args.data_corte}'. Use formato DD/MM/YYYY.")
        sys.exit(1)

    print(f"=== nps-acao-contato | data de corte: {cutoff_dt.strftime('%d/%m/%Y')} ===\n")
    date_tag = cutoff_dt.strftime("%m%d")
    year_tag = cutoff_dt.strftime("%Y")
    out_xlsx = os.path.join(OUTPUT_DIR, f"reclamacoes_{date_tag}_{year_tag}.xlsx")
    out_meta = os.path.join(OUTPUT_DIR, f"reclamacoes_{date_tag}_{year_tag}_meta.json")

    # 1. Carregar dados NPS
    print("[1/5] Carregando dados NPS do Google Sheets...")
    nps_rows = load_nps_from_sheets()
    if nps_rows:
        print(f"  Sheets: {len(nps_rows)} linhas carregadas.")
    else:
        print("  Sheets indisponível — tentando cache local...")
        nps_rows = load_nps_from_cache()
        if nps_rows:
            print(f"  Cache: {len(nps_rows)} linhas carregadas.")
        else:
            print("[ERRO] Nem Sheets nem cache disponíveis. Abortando.")
            sys.exit(1)

    # 2. Filtrar detratores
    print(f"\n[2/5] Filtrando detratores >= {cutoff_dt.strftime('%d/%m/%Y')}...")
    final_rows, keyword_filter_applied = filter_detractors(nps_rows, cutoff_dt)

    if not final_rows:
        print("[!] Nenhum detrator encontrado para o período informado. Arquivo não gerado.")
        sys.exit(0)

    print(f"  TOTAL FINAL: {len(final_rows)} registros")

    # 3. Coletar driver_uuids
    print("\n[3/5] Coletando driver_uuids para cruzamento PII...")
    driver_uuids = set()
    for row in final_rows:
        if len(row) > IDX["driver_uuid"]:
            uid = str(row[IDX["driver_uuid"]]).strip()
            if uid and uid.lower() not in ("", "none"):
                driver_uuids.add(uid)
    print(f"  Driver UUIDs identificados: {len(driver_uuids)}")

    # 4. Consultar Databricks PII
    print("\n[4/5] Consultando Databricks PII (nome, CPF, telefone)...")
    pii_map = query_pii(driver_uuids)

    # 5. Montar e salvar
    print("\n[5/5] Gerando planilha e metadados...")
    output_rows = build_output_rows(final_rows, pii_map)
    write_excel(output_rows, out_xlsx)
    print(f"  Planilha salva: {out_xlsx}")
    meta = write_meta(output_rows, keyword_filter_applied, out_meta)
    print(f"  Metadados salvos: {out_meta}")
    # Resumo final
    print("\n" + "=" * 60)
    print("✅ CONCLUÍDO!")
    print(f"  Arquivo       : {out_xlsx}")
    print(f"  Total casos   : {meta['total_detratores_filtrados']}")
    print(f"  Com telefone  : {meta['com_telefone']}")
    print(f"  Período       : {meta['data_minima']} → {meta['data_maxima']}")
    print(f"  Filtro keyword: {'sim' if keyword_filter_applied else 'não (fallback)'}")
    print("=" * 60)
    print(json.dumps(meta, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    main()
